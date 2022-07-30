import openpyxl as op
import tkinter as tk
import tkinter.filedialog as fd
import os
import sys
import string

d = dict(enumerate(string.ascii_uppercase, 1))
again = True

base = tk.Tk()
# base.withdraw()
base.geometry('400x120')
base.title("Grade Viwer")
var = tk.IntVar()

currdir = os.getcwd()

path = ''


def OpenFile():
    base.lift()
    base.attributes('-topmost', True)
    base.after_idle(base.attributes, '-topmost', False)

    filePath = fd.askopenfilename(parent=base, initialdir=currdir,
                                  title='Please select an excel file.',
                                  filetypes=[
                                     ("Excel Files", ".xlsx")
                                  ])

    global path
    path = filePath

    var.set(1)


label = tk.Label(base, text="Click the button to select an Excel File",
                 font=('Segoi_UI_Light 16'))
label.pack(pady=10)

browseButton = tk.Button(text="Browse Files", font=('Segoi_UI-Light 12'),
                         command=lambda: OpenFile())
browseButton.pack(pady=20)

browseButton.wait_variable(var)

while again is True:

    filePath = path

    if filePath != '':
        print("The chosen file is :", filePath + "\n")
        again = False
        break
    else:
        a = input("No file selected. Press 'n' if you want to exit. Press 'y' "
                  + "after selecting a file to continue : ")

    if a == "n":
        sys.exit()

base.withdraw()

print("\n------------------------------\n")

os.system("pause")

subNum = int(input("\nEnter Subject Code : "))
wb = op.load_workbook(path)
sheet_obj = wb.active

firstRow = sheet_obj.min_row
firstCol = sheet_obj.min_column
maxRow = sheet_obj.max_row
maxCol = sheet_obj.max_column

cell_obj = sheet_obj[str(d[firstCol]) + str(firstRow):
                     str(d[maxCol]) + str(maxRow)]

subNumCheck = sheet_obj['B1':'B' + str(maxRow)]

newWB = op.Workbook()
ws = newWB.active
ws.title = "Marks for Sub. Code " + str(subNum)

ws['A1'] = "Name"
ws['B1'] = "Marks"
ws['C1'] = "Grade"

h = 3

if subNum in iter(subNumCheck):
    pass
else:
    print("Subject Code doesn't exist. Empty File will be generated.")

for cell1, cell2, cell3 in cell_obj:
    if cell2.value == subNum:
        cell2 = sheet_obj.cell(row=cell2.row + 1, column=cell2.column)
        cell3 = sheet_obj.cell(row=cell3.row + 1, column=cell3.column)
        ws.cell(row=h, column=1).value = cell1.value
        ws.cell(row=h, column=2).value = cell2.value
        ws.cell(row=h, column=3).value = cell3.value

        h += 2


savePath = fd.askdirectory(initialdir=currdir, title="Select location to save")

if savePath:
    newWB.save(str(savePath) + "/Student Marks.xlsx")
    print("New Excel file created! \nFile Located in :", savePath)
else:
    print("Folder not selected, saving to desktop.")
    newWB.save(os.path.expanduser('~') + "/desktop" + "/Student Marks.xlsx")

print("\n------------------------------\n")

os.system("pause")
